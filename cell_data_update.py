import openpyxl

word_find = '3600' 
word_replace = 'updated'

file_path = 'updated_Ref_Input File.xlsx'



detected_loc = []
def wordfinder(searchString):
    for i in range(1, sheet_obj_0.max_row + 1):
        for j in range(1, sheet_obj_0.max_column + 1):
            if(searchString == str(sheet_obj_0.cell(i,j).value)):
                detected_loc.append((i,j))
    if(len(detected_loc)==0):
        print("Data not found")
    else:
        print("Data found")


        




def wordreplace(word_replace):
    wb = openpyxl.Workbook()
    sheet = wb.active
    for i in range(1, sheet_obj_0.max_row + 1):
        for j in range(1, sheet_obj_0.max_column + 1):
            c1 = sheet.cell(row = i, column = j)
            if((i,j) in detected_loc):
                c1.value = word_replace
            else:
                c1.value = sheet_obj_0.cell(i,j).value
    wb.save(file_path)
    print("DONE......")

    



 

wb_obj_0 = openpyxl.load_workbook(file_path)
sheet_obj_0 = wb_obj_0.active






print(word_find + " ---> " + word_replace)
wordfinder(word_find)
wordreplace(word_replace)




  
